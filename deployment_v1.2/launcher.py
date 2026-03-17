"""
Transfer Agent Launcher
Double-click → opens browser → shows live progress → DONE
Uses only stdlib + openpyxl (already in .venv)
"""
import os
import sys
import queue
import sqlite3
import threading
import webbrowser
from io import BytesIO
from datetime import datetime, date, time
from http.server import BaseHTTPRequestHandler, HTTPServer

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ─── Config ───────────────────────────────────────────────────────────────────
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH   = os.path.join(BASE_DIR, "LOGBOOK.xlsx")
DB_PATH     = os.path.join(BASE_DIR, "logbook.db")
SHEET_NAME  = "Logbook"
DATE_FROM   = date(2020, 1, 28)
DATE_TO     = date.today()
PORT        = 7654

# ─── Shared progress queue ────────────────────────────────────────────────────
_queue: queue.Queue = queue.Queue()

SOURCE_MAP = {
    "datetime": "A", "event": "C", "place": "D",
    "me_rev_c": "E", "main_flmtr": "F", "dg_in_flmtr": "G",
    "dg_out_flmtr": "H", "blr_flmtr": "I", "cyl_oil_count": "J",
    "me_pwrmtr": "K", "me_hrs": "CA", "dg1_hrs": "CB", "dg2_hrs": "CC",
    "dg3_hrs": "CD", "boiler_hrs": "CE", "hfo_bnkr": "BJ", "do_bnkr": "BL",
    "me_sys_bnkr": "BM", "me_cyl_bnkr": "BN", "dg_sys_bnkr": "BO",
    "me_hfo_cor_cons": "AK", "me_do_cor_cons": "AS",
    "me_sys_cor_cons": "AZ", "me_cyl_cor_cons": "BA", "dg_sys_cor_cons": "BB",
}
INTEGER_KEYS = {"me_rev_c","main_flmtr","dg_in_flmtr","dg_out_flmtr","blr_flmtr","cyl_oil_count"}

# ─── Parsers ──────────────────────────────────────────────────────────────────
def parse_datetime_cell(v):
    """Extract (date_obj, time_str) from a single cell containing both date and time."""
    if v is None or v == "": return None, ""
    if isinstance(v, datetime): return v.date(), v.strftime("%H:%M")
    if isinstance(v, date): return v, ""
    if isinstance(v, time): return None, v.strftime("%H:%M")
    if isinstance(v, str):
        s = v.strip()
        if not s: return None, ""
        for fmt in ("%d-%m-%y %H:%M","%d.%m.%Y %H:%M","%Y-%m-%d %H:%M",
                     "%d/%m/%Y %H:%M","%d/%m/%y %H:%M",
                     "%d-%m-%y %H:%M:%S","%d.%m.%Y %H:%M:%S","%Y-%m-%d %H:%M:%S"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.date(), dt.strftime("%H:%M")
            except ValueError: pass
        for fmt in ("%d-%m-%y","%d.%m.%Y","%Y-%m-%d","%d/%m/%Y","%d/%m/%y"):
            try: return datetime.strptime(s, fmt).date(), ""
            except ValueError: pass
        return None, ""
    return None, ""

def parse_num_cell(v, is_int=False):
    if v is None or v == "": return None
    try: n = float(v)
    except: return None
    return int(round(n)) if is_int else round(n,2)

# ─── Core transfer logic ──────────────────────────────────────────────────────
def run_transfer():
    try:
        _queue.put(("status", "Checking files..."))

        if not os.path.isfile(XLSX_PATH):
            _queue.put(("error", f"LOGBOOK.xlsx not found in {BASE_DIR}"))
            return
        if not os.path.isfile(DB_PATH):
            _queue.put(("error", f"logbook.db not found in {BASE_DIR}"))
            return

        _queue.put(("status", "Reading source Excel..."))
        col_idx = {k: column_index_from_string(c) for k,c in SOURCE_MAP.items()}
        wb = load_workbook(filename=XLSX_PATH, data_only=True, read_only=True)
        sheet = wb[SHEET_NAME]

        all_records = []
        for ri, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            n = len(row)
            def g(key):
                i = col_idx[key]
                return row[i-1] if i <= n else None
            pd_, pt_ = parse_datetime_cell(g("datetime"))
            if pd_ is None: continue
            rec = {
                "date": pd_.strftime("%d-%m-%y"), "_date_obj": pd_,
                "time": pt_,
                "event": str(g("event") or "").strip(),
                "place": str(g("place") or "").strip(),
                "me_fo_set":"HFO","dg_fo_set":"HFO","blr_fo_set":"DO",
                "_source_row": ri,
            }
            for key in SOURCE_MAP:
                if key in {"datetime","event","place"}: continue
                rec[key] = parse_num_cell(g(key), is_int=(key in INTEGER_KEYS))
            all_records.append(rec)
        wb.close()

        selected = sorted(
            [r for r in all_records if DATE_FROM <= r["_date_obj"] <= DATE_TO],
            key=lambda r: (r["_date_obj"], r.get("time",""), r["_source_row"]),
        )
        total = len(selected)
        _queue.put(("found", str(total)))
        if total == 0:
            _queue.put(("error", "No rows found in the selected date range."))
            return

        # ── DB insert ──
        conn = sqlite3.connect(DB_PATH)
        try:
            if not conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='events'").fetchone():
                raise RuntimeError("Table 'events' not found in database.")
            cols_meta = conn.execute("PRAGMA table_info(events)").fetchall()
            event_columns = [r[1] for r in cols_meta]

            conn.execute("BEGIN")
            # ensure seed
            if not conn.execute("SELECT * FROM events WHERE id=1").fetchone():
                sv = {"id":1,"date":"28-01-26","time":"01:10","event":"NOON","place":"SEED",
                      "me_rev_c":1,"main_flmtr":1,"dg_in_flmtr":1,"dg_out_flmtr":1,
                      "blr_flmtr":1,"cyl_oil_count":1,"me_pwrmtr":1,
                      "me_hrs":1.0,"dg1_hrs":1.0,"dg2_hrs":1.0,"dg3_hrs":1.0,
                      "boiler_hrs":1.0,"dg1_mwh":1.0,"dg2_mwh":1.0,"dg3_mwh":1.0,
                      "sox_co2":1.0,"me_fo_set":"HFO","dg_fo_set":"HFO","blr_fo_set":"DO"}
                ik = [k for k in sv if k in event_columns]
                conn.execute(f"INSERT INTO events ({','.join(ik)}) VALUES ({','.join(['?']*len(ik))})",
                             [sv[k] for k in ik])
            conn.execute("DELETE FROM events WHERE id<>1")
            try: conn.execute("DELETE FROM sqlite_sequence WHERE name='events'")
            except: pass

            insertable = [k for k in [
                "id","date","time","event","place","me_rev_c","main_flmtr",
                "dg_in_flmtr","dg_out_flmtr","blr_flmtr","cyl_oil_count","me_pwrmtr",
                "me_hrs","dg1_hrs","dg2_hrs","dg3_hrs","boiler_hrs","hfo_bnkr","do_bnkr",
                "me_sys_bnkr","me_cyl_bnkr","dg_sys_bnkr","me_hfo_cor_cons","me_do_cor_cons",
                "me_sys_cor_cons","me_cyl_cor_cons","dg_sys_cor_cons",
                "me_fo_set","dg_fo_set","blr_fo_set",
            ] if k in event_columns]

            BATCH = 200
            for i, src in enumerate(selected, start=1):
                payload = {k: src.get(k) for k in insertable if k != "id"}
                payload["id"] = i + 1
                conn.execute(
                    f"INSERT INTO events ({','.join(insertable)}) VALUES ({','.join(['?']*len(insertable))})",
                    [payload[k] for k in insertable],
                )
                if i % BATCH == 0:
                    _queue.put(("progress", str(i)))

            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

        _queue.put(("progress", str(total)))
        _queue.put(("done", str(total)))

    except Exception as e:
        _queue.put(("error", str(e)))


# ─── HTML page ────────────────────────────────────────────────────────────────
HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Transfer Agent</title>
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{background:#0d1117;color:#e6edf3;font-family:'Segoe UI',sans-serif;
       display:flex;align-items:center;justify-content:center;min-height:100vh}
  .card{background:#161b22;border:1px solid #30363d;border-radius:12px;
        padding:48px 56px;min-width:420px;max-width:560px;text-align:center;
        box-shadow:0 8px 32px rgba(0,0,0,.4)}
  h1{font-size:1.5rem;font-weight:600;margin-bottom:8px;color:#f0f6fc}
  .sub{font-size:.85rem;color:#8b949e;margin-bottom:32px}
  .icon{font-size:3rem;margin-bottom:16px;line-height:1}
  .status{font-size:1rem;color:#8b949e;margin-bottom:24px;min-height:1.4em}
  .bar-wrap{background:#21262d;border-radius:999px;height:8px;margin-bottom:10px;overflow:hidden}
  .bar{background:linear-gradient(90deg,#1f6feb,#58a6ff);height:100%;width:0%;
       transition:width .4s ease;border-radius:999px}
  .count{font-size:.8rem;color:#484f58;margin-bottom:32px}
  .done-msg{font-size:1.3rem;font-weight:700;color:#3fb950;margin:16px 0 8px}
  .done-sub{font-size:.85rem;color:#8b949e}
  .error-msg{color:#f85149;font-size:.9rem;margin-top:16px}
  .btn{display:inline-block;margin-top:28px;padding:10px 32px;
       background:#238636;color:#fff;border:none;border-radius:6px;
       font-size:.95rem;cursor:pointer;transition:background .2s}
  .btn:hover{background:#2ea043}
</style>
</head>
<body>
<div class="card">
  <div class="icon" id="icon">⚙️</div>
  <h1>Transfer Agent</h1>
  <div class="sub" id="range">LOGBOOK.xlsx &rarr; logbook.db</div>
  <div class="status" id="status">Connecting...</div>
  <div class="bar-wrap"><div class="bar" id="bar"></div></div>
  <div class="count" id="count">&nbsp;</div>
  <div id="extra"></div>
</div>
<script>
const status = document.getElementById('status');
const bar    = document.getElementById('bar');
const count  = document.getElementById('count');
const icon   = document.getElementById('icon');
const extra  = document.getElementById('extra');
let total = 0;

const es = new EventSource('/events');
es.onmessage = e => {
  const [type, val] = e.data.split('|', 2);
  if (type === 'status') {
    status.textContent = val;
  } else if (type === 'found') {
    total = parseInt(val);
    status.textContent = `Transferring ${total.toLocaleString()} rows...`;
    count.textContent  = `0 / ${total.toLocaleString()}`;
  } else if (type === 'progress') {
    const n = parseInt(val);
    const pct = total > 0 ? Math.round(n/total*100) : 0;
    bar.style.width = pct + '%';
    count.textContent = `${n.toLocaleString()} / ${total.toLocaleString()} — ${pct}%`;
  } else if (type === 'done') {
    const n = parseInt(val);
    bar.style.width = '100%';
    icon.textContent = '✅';
    status.textContent = '';
    count.textContent = '';
    extra.innerHTML = `<div class="done-msg">Transfer Complete</div>
      <div class="done-sub">${n.toLocaleString()} rows imported &mdash; seed preserved</div>
      <button class="btn" onclick="window.close()">Close</button>`;
    es.close();
  } else if (type === 'error') {
    icon.textContent = '❌';
    status.textContent = '';
    extra.innerHTML = `<div class="error-msg">${val}</div>
      <button class="btn" style="background:#b62324" onclick="window.close()">Close</button>`;
    es.close();
  }
};
es.onerror = () => { status.textContent = 'Connection lost.'; };
</script>
</body>
</html>
"""

# ─── HTTP server ──────────────────────────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a): pass  # silence access log

    def do_GET(self):
        if self.path == "/":
            body = HTML.encode()
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        elif self.path == "/events":
            self.send_response(200)
            self.send_header("Content-Type", "text/event-stream")
            self.send_header("Cache-Control", "no-cache")
            self.send_header("X-Accel-Buffering", "no")
            self.end_headers()
            while True:
                try:
                    evt_type, val = _queue.get(timeout=30)
                    msg = f"data: {evt_type}|{val}\n\n"
                    self.wfile.write(msg.encode())
                    self.wfile.flush()
                    if evt_type in ("done", "error"):
                        break
                except queue.Empty:
                    # keep-alive comment
                    try:
                        self.wfile.write(b": ping\n\n")
                        self.wfile.flush()
                    except BrokenPipeError:
                        break
                except BrokenPipeError:
                    break
        else:
            self.send_response(404)
            self.end_headers()


def start_server():
    server = HTTPServer(("127.0.0.1", PORT), Handler)
    server.serve_forever()


# ─── Entry point ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Start HTTP server in background thread
    t = threading.Thread(target=start_server, daemon=True)
    t.start()

    # Open browser after a short delay
    threading.Timer(0.6, lambda: webbrowser.open(f"http://127.0.0.1:{PORT}/")).start()

    # Run transfer in another thread so main thread stays alive
    worker = threading.Thread(target=run_transfer)
    worker.start()
    worker.join()

    # Keep server alive briefly so browser can receive final event + close
    import time as _time
    _time.sleep(3)
    sys.exit(0)
