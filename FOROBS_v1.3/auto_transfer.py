"""
Auto Transfer: reads LOGBOOK.xlsx → logbook.db
Range: 2020-01-28 to today
"""
import os
import sqlite3
from io import BytesIO
from datetime import datetime, date, time

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

SOURCE_MAP = {
    "datetime": "A",
    "event": "C",
    "place": "D",
    "me_rev_c": "E",
    "main_flmtr": "F",
    "dg_in_flmtr": "G",
    "dg_out_flmtr": "H",
    "blr_flmtr": "I",
    "cyl_oil_count": "J",
    "me_pwrmtr": "K",
    "me_hrs": "CA",
    "dg1_hrs": "CB",
    "dg2_hrs": "CC",
    "dg3_hrs": "CD",
    "boiler_hrs": "CE",
    "hfo_bnkr": "BJ",
    "do_bnkr": "BL",
    "me_sys_bnkr": "BM",
    "me_cyl_bnkr": "BN",
    "dg_sys_bnkr": "BO",
    "me_hfo_cor_cons": "AK",
    "me_do_cor_cons": "AS",
    "me_sys_cor_cons": "AZ",
    "me_cyl_cor_cons": "BA",
    "dg_sys_cor_cons": "BB",
}

INTEGER_KEYS = {
    "me_rev_c", "main_flmtr", "dg_in_flmtr",
    "dg_out_flmtr", "blr_flmtr", "cyl_oil_count",
}


def parse_datetime_cell(value):
    """Extract (date_obj, time_str) from a single cell that may contain both date and time."""
    if value is None or value == "":
        return None, ""
    if isinstance(value, datetime):
        return value.date(), value.strftime("%H:%M")
    if isinstance(value, date):
        return value, ""
    if isinstance(value, time):
        return None, value.strftime("%H:%M")
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None, ""
        # Try combined date+time formats first
        for fmt in ("%d-%m-%y %H:%M", "%d.%m.%Y %H:%M", "%Y-%m-%d %H:%M",
                     "%d/%m/%Y %H:%M", "%d/%m/%y %H:%M",
                     "%d-%m-%y %H:%M:%S", "%d.%m.%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                dt = datetime.strptime(text, fmt)
                return dt.date(), dt.strftime("%H:%M")
            except ValueError:
                continue
        # Try date-only formats
        for fmt in ("%d-%m-%y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(text, fmt).date(), ""
            except ValueError:
                continue
        return None, ""
    return None, ""


def parse_num_cell(value, is_int=False):
    if value is None or value == "":
        return None
    try:
        num = float(value)
    except (ValueError, TypeError):
        return None
    if is_int:
        return int(round(num))
    return round(num, 2)


def read_source_records(xlsx_path, sheet_name):
    # Build a 1-based column-index lookup for each key we care about
    col_indices = {key: column_index_from_string(col) for key, col in SOURCE_MAP.items()}

    workbook = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
    sheet = workbook[sheet_name]

    records = []
    for row_idx, row_cells in enumerate(sheet.iter_rows(values_only=True), start=1):
        max_col = len(row_cells)

        def get(col_letter):
            idx = col_indices.get(col_letter) or column_index_from_string(col_letter)
            if idx > max_col:
                return None
            return row_cells[idx - 1]

        # Use the column index directly for each key
        def getk(key):
            idx = col_indices[key]
            if idx > max_col:
                return None
            return row_cells[idx - 1]

        raw_dt = getk("datetime")
        parsed_date, parsed_time = parse_datetime_cell(raw_dt)
        if parsed_date is None:
            continue

        record = {
            "date": parsed_date.strftime("%d-%m-%y"),
            "_date_obj": parsed_date,
            "time": parsed_time,
            "event": str(getk("event") or "").strip(),
            "place": str(getk("place") or "").strip(),
            "me_fo_set": "HFO",
            "dg_fo_set": "HFO",
            "blr_fo_set": "DO",
            "_source_row": row_idx,
        }

        for key in SOURCE_MAP:
            if key in {"datetime", "event", "place"}:
                continue
            record[key] = parse_num_cell(getk(key), is_int=(key in INTEGER_KEYS))

        records.append(record)

    workbook.close()
    return records


def ensure_seed(conn, event_columns):
    row = conn.execute("SELECT * FROM events WHERE id = 1").fetchone()
    if row is not None:
        return

    seed_values = {
        "id": 1, "date": "28-01-26", "time": "01:10", "event": "NOON", "place": "SEED",
        "me_rev_c": 1, "main_flmtr": 1, "dg_in_flmtr": 1, "dg_out_flmtr": 1,
        "blr_flmtr": 1, "cyl_oil_count": 1, "me_pwrmtr": 1,
        "me_hrs": 1.0, "dg1_hrs": 1.0, "dg2_hrs": 1.0, "dg3_hrs": 1.0,
        "boiler_hrs": 1.0, "dg1_mwh": 1.0, "dg2_mwh": 1.0, "dg3_mwh": 1.0,
        "sox_co2": 1.0, "me_fo_set": "HFO", "dg_fo_set": "HFO", "blr_fo_set": "DO",
    }

    insert_keys = [k for k in seed_values.keys() if k in event_columns]
    cols = ", ".join(insert_keys)
    vals = [seed_values[k] for k in insert_keys]
    placeholders = ", ".join(["?" for _ in vals])
    conn.execute(f"INSERT INTO events ({cols}) VALUES ({placeholders})", vals)


def transfer_records(db_path, records):
    conn = sqlite3.connect(db_path)
    try:
        table_exists = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='events'"
        ).fetchone()
        if not table_exists:
            raise RuntimeError("Table 'events' not found in selected database.")

        cols_meta = conn.execute("PRAGMA table_info(events)").fetchall()
        event_columns = [row[1] for row in cols_meta]

        conn.execute("BEGIN")
        ensure_seed(conn, event_columns)
        conn.execute("DELETE FROM events WHERE id <> 1")
        try:
            conn.execute("DELETE FROM sqlite_sequence WHERE name='events'")
        except sqlite3.Error:
            pass

        insertable = [
            "id", "date", "time", "event", "place",
            "me_rev_c", "main_flmtr", "dg_in_flmtr", "dg_out_flmtr", "blr_flmtr",
            "cyl_oil_count", "me_pwrmtr", "me_hrs", "dg1_hrs", "dg2_hrs", "dg3_hrs",
            "boiler_hrs", "hfo_bnkr", "do_bnkr", "me_sys_bnkr", "me_cyl_bnkr",
            "dg_sys_bnkr", "me_hfo_cor_cons", "me_do_cor_cons", "me_sys_cor_cons",
            "me_cyl_cor_cons", "dg_sys_cor_cons", "me_fo_set", "dg_fo_set", "blr_fo_set",
        ]
        insertable = [k for k in insertable if k in event_columns]

        next_id = 2
        for src in records:
            payload = {k: src.get(k) for k in insertable if k != "id"}
            payload["id"] = next_id
            cols = ", ".join(insertable)
            vals = [payload[k] for k in insertable]
            placeholders = ", ".join(["?" for _ in vals])
            conn.execute(f"INSERT INTO events ({cols}) VALUES ({placeholders})", vals)
            next_id += 1

        conn.commit()
        return len(records)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    base = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(base, "LOGBOOK.xlsx")
    db_path = os.path.join(base, "logbook.db")
    sheet_name = "Logbook"

    date_from = date(2020, 1, 28)
    date_to = date.today()

    print(f"Source : {xlsx_path}")
    print(f"Target : {db_path}")
    print(f"Range  : {date_from} → {date_to}")

    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"Source not found: {xlsx_path}")
    if not os.path.isfile(db_path):
        raise FileNotFoundError(f"Database not found: {db_path}")

    print("Reading source records...")
    all_records = read_source_records(xlsx_path, sheet_name)
    print(f"  Total rows in source: {len(all_records)}")

    selected = sorted(
        [r for r in all_records if date_from <= r["_date_obj"] <= date_to],
        key=lambda r: (r["_date_obj"], r.get("time", ""), r["_source_row"]),
    )
    print(f"  Rows in selected range: {len(selected)}")

    if not selected:
        print("No rows to transfer. Exiting.")
    else:
        transferred = transfer_records(db_path, selected)
        print(f"Done. {transferred} rows imported into {db_path}.")
