import os
import sqlite3
from io import BytesIO
from datetime import datetime, date, time

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


st.set_page_config(page_title="Database Transfer Agent", layout="wide")


SOURCE_MAP = {
    "datetime": "A",
    "event": "C",
    "place": "D",
    "me_rev_c": "E",
    "main_flmtr": "F",
    "dg_in_flmtr": "G",
    "dg_out_flmtr": "H",
    "blr_flmtr": "I",
    "cyl_oil_count": "J",
    "me_pwrmtr": "K",
    "me_hrs": "CA",
    "dg1_hrs": "CB",
    "dg2_hrs": "CC",
    "dg3_hrs": "CD",
    "boiler_hrs": "CE",
    "hfo_bnkr": "BJ",
    "do_bnkr": "BL",
    "me_sys_bnkr": "BM",
    "me_cyl_bnkr": "BN",
    "dg_sys_bnkr": "BO",
    "me_hfo_cor_cons": "AK",
    "me_do_cor_cons": "AS",
    "me_sys_cor_cons": "AZ",
    "me_cyl_cor_cons": "BA",
    "dg_sys_cor_cons": "BB",
}


INTEGER_KEYS = {
    "me_rev_c",
    "main_flmtr",
    "dg_in_flmtr",
    "dg_out_flmtr",
    "blr_flmtr",
    "cyl_oil_count",
}


def parse_datetime_cell(value):
    """Extract (date_obj, time_str) from a single cell that may contain both date and time."""
    if value is None or value == "":
        return None, ""
    if isinstance(value, datetime):
        return value.date(), value.strftime("%H:%M")
    if isinstance(value, date):
        return value, ""
    if isinstance(value, time):
        return None, value.strftime("%H:%M")
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None, ""
        # Try combined date+time formats first
        for fmt in ("%d-%m-%y %H:%M", "%d.%m.%Y %H:%M", "%Y-%m-%d %H:%M",
                     "%d/%m/%Y %H:%M", "%d/%m/%y %H:%M",
                     "%d-%m-%y %H:%M:%S", "%d.%m.%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
            try:
                dt = datetime.strptime(text, fmt)
                return dt.date(), dt.strftime("%H:%M")
            except ValueError:
                continue
        # Try date-only formats
        for fmt in ("%d-%m-%y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(text, fmt).date(), ""
            except ValueError:
                continue
        return None, ""
    return None, ""


def parse_num_cell(value, is_int=False):
    if value is None or value == "":
        return None
    try:
        num = float(value)
    except (ValueError, TypeError):
        return None
    if is_int:
        return int(round(num))
    return round(num, 2)


def get_cell_value(sheet, row_idx, col_letter):
    col_idx = column_index_from_string(col_letter)
    return sheet.cell(row=row_idx, column=col_idx).value


def read_source_records(xlsx_bytes, sheet_name):
    workbook = load_workbook(filename=xlsx_bytes, data_only=True, read_only=True)
    sheet = workbook[sheet_name]

    records = []
    for row_idx in range(1, sheet.max_row + 1):
        raw_dt = get_cell_value(sheet, row_idx, SOURCE_MAP["datetime"])
        parsed_date, parsed_time = parse_datetime_cell(raw_dt)
        if parsed_date is None:
            continue

        record = {
            "date": parsed_date.strftime("%d-%m-%y"),
            "_date_obj": parsed_date,
            "time": parsed_time,
            "event": str(get_cell_value(sheet, row_idx, SOURCE_MAP["event"]) or "").strip(),
            "place": str(get_cell_value(sheet, row_idx, SOURCE_MAP["place"]) or "").strip(),
            "me_fo_set": "HFO",
            "dg_fo_set": "HFO",
            "blr_fo_set": "DO",
            "_source_row": row_idx,
        }

        for key, col in SOURCE_MAP.items():
            if key in {"datetime", "event", "place"}:
                continue
            record[key] = parse_num_cell(get_cell_value(sheet, row_idx, col), is_int=(key in INTEGER_KEYS))

        records.append(record)

    return records


def ensure_seed(conn, event_columns):
    row = conn.execute("SELECT * FROM events WHERE id = 1").fetchone()
    if row is not None:
        return

    seed_values = {
        "id": 1,
        "date": "28-01-26",
        "time": "01:10",
        "event": "NOON",
        "place": "SEED",
        "me_rev_c": 1,
        "main_flmtr": 1,
        "dg_in_flmtr": 1,
        "dg_out_flmtr": 1,
        "blr_flmtr": 1,
        "cyl_oil_count": 1,
        "me_pwrmtr": 1,
        "me_hrs": 1.0,
        "dg1_hrs": 1.0,
        "dg2_hrs": 1.0,
        "dg3_hrs": 1.0,
        "boiler_hrs": 1.0,
        "dg1_mwh": 1.0,
        "dg2_mwh": 1.0,
        "dg3_mwh": 1.0,
        "sox_co2": 1.0,
        "me_fo_set": "HFO",
        "dg_fo_set": "HFO",
        "blr_fo_set": "DO",
    }

    insert_keys = [k for k in seed_values.keys() if k in event_columns]
    cols = ", ".join(insert_keys)
    vals = [seed_values[k] for k in insert_keys]
    placeholders = ", ".join(["?" for _ in vals])
    conn.execute(f"INSERT INTO events ({cols}) VALUES ({placeholders})", vals)


def transfer_records(db_path, records):
    conn = sqlite3.connect(db_path)
    try:
        table_exists = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='events'"
        ).fetchone()
        if not table_exists:
            raise RuntimeError("Table 'events' not found in selected database.")

        cols_meta = conn.execute("PRAGMA table_info(events)").fetchall()
        event_columns = [row[1] for row in cols_meta]

        conn.execute("BEGIN")
        ensure_seed(conn, event_columns)
        conn.execute("DELETE FROM events WHERE id <> 1")
        try:
            conn.execute("DELETE FROM sqlite_sequence WHERE name='events'")
        except sqlite3.Error:
            pass

        insertable = [
            "id",
            "date",
            "time",
            "event",
            "place",
            "me_rev_c",
            "main_flmtr",
            "dg_in_flmtr",
            "dg_out_flmtr",
            "blr_flmtr",
            "cyl_oil_count",
            "me_pwrmtr",
            "me_hrs",
            "dg1_hrs",
            "dg2_hrs",
            "dg3_hrs",
            "boiler_hrs",
            "hfo_bnkr",
            "do_bnkr",
            "me_sys_bnkr",
            "me_cyl_bnkr",
            "dg_sys_bnkr",
            "me_hfo_cor_cons",
            "me_do_cor_cons",
            "me_sys_cor_cons",
            "me_cyl_cor_cons",
            "dg_sys_cor_cons",
            "me_fo_set",
            "dg_fo_set",
            "blr_fo_set",
        ]
        insertable = [k for k in insertable if k in event_columns]

        next_id = 2
        for src in records:
            payload = {k: src.get(k) for k in insertable if k != "id"}
            payload["id"] = next_id
            cols = ", ".join(insertable)
            vals = [payload[k] for k in insertable]
            placeholders = ", ".join(["?" for _ in vals])
            conn.execute(f"INSERT INTO events ({cols}) VALUES ({placeholders})", vals)
            next_id += 1

        conn.commit()
        return len(records)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def list_db_files(base_dir):
    return sorted([
        f for f in os.listdir(base_dir)
        if f.lower().endswith(".db") and os.path.isfile(os.path.join(base_dir, f))
    ])


st.title("Transfer Agent")
st.caption("Simple import menu: Input DB → Source Excel → Date range → Import")

workspace_dir = os.path.dirname(os.path.abspath(__file__))
db_files = list_db_files(workspace_dir)
local_source_path = os.path.join(workspace_dir, "OLD_LOGBOOK.xlsx")

st.subheader("1) Input Database")
if db_files:
    default_idx = db_files.index("logbook.db") if "logbook.db" in db_files else 0
    selected_db_name = st.selectbox("Input DB file", db_files, index=default_idx)
else:
    selected_db_name = ""
    st.warning("No .db files found in current folder.")

custom_db_path = st.text_input("Or full DB path", value="")
db_path = custom_db_path.strip() if custom_db_path.strip() else (os.path.join(workspace_dir, selected_db_name) if selected_db_name else "")

st.subheader("2) Source Excel")
source_mode = st.radio(
    "Source mode",
    options=["Use local OLD_LOGBOOK.xlsx", "Upload source file"],
    horizontal=True,
)

source_bytes = None
if source_mode == "Use local OLD_LOGBOOK.xlsx":
    if os.path.isfile(local_source_path):
        st.success(f"Source: {os.path.basename(local_source_path)}")
        with open(local_source_path, "rb") as source_fp:
            source_bytes = source_fp.read()
    else:
        st.error("Local source file OLD_LOGBOOK.xlsx not found.")
else:
    source_file = st.file_uploader("Upload source (.xlsx / .xlsm)", type=["xlsx", "xlsm"])
    source_bytes = source_file.read() if source_file else None

if not db_path:
    st.info("Select input DB first.")
    st.stop()

if not source_bytes:
    st.info("Select source Excel file.")
    st.stop()

sheet_name = "Logbook"
try:
    wb_preview = load_workbook(filename=BytesIO(source_bytes), data_only=True, read_only=True)
    if sheet_name not in wb_preview.sheetnames:
        st.error(f"Worksheet '{sheet_name}' not found in source workbook.")
        st.stop()
except Exception as ex:
    st.error(f"Cannot open source workbook: {ex}")
    st.stop()

records = read_source_records(BytesIO(source_bytes), sheet_name)
if not records:
    st.error("No valid source rows found (DATE column A is required).")
    st.stop()

records = sorted(records, key=lambda r: (r["_date_obj"], r.get("time", ""), r["_source_row"]))
min_d = min(r["_date_obj"] for r in records)
max_d = max(r["_date_obj"] for r in records)

st.subheader("3) Import Range")
from_to = st.date_input(
    "Date range",
    value=(min_d, max_d),
    min_value=min_d,
    max_value=max_d,
)

if isinstance(from_to, tuple) and len(from_to) == 2:
    date_from, date_to = from_to
else:
    date_from, date_to = min_d, max_d

if date_from > date_to:
    date_from, date_to = date_to, date_from

selected_records = [r for r in records if date_from <= r["_date_obj"] <= date_to]

st.subheader("4) Summary")
st.write(f"Rows in source: {len(records)}")
st.write(f"Selected range: {date_from} → {date_to}")
st.write(f"Rows to import: {len(selected_records)}")

preview_cols = ["date", "time", "event", "place", "me_rev_c", "main_flmtr", "me_hrs", "hfo_bnkr", "me_hfo_cor_cons"]
preview_df = pd.DataFrame([{k: r.get(k) for k in preview_cols} for r in selected_records[:100]])
st.dataframe(preview_df, hide_index=True, use_container_width=True)

st.warning("Import will DELETE all current events except seed (ID=1), then load selected rows.")
if st.button("IMPORT", type="primary", use_container_width=True, disabled=(len(selected_records) == 0)):
    try:
        transferred = transfer_records(db_path, selected_records)
        st.success(f"Import complete: {transferred} rows imported. Seed (ID=1) preserved.")
    except Exception as ex:
        st.error(f"Import failed: {ex}")
